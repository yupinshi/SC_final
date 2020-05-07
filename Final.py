'''
EN.540.635 Software Carpentry
Final Project
Yupin Shi
'''

from PIL import Image, ImageDraw
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

'''
Define global constants
'''
length = 0.015
height_pas = 0.0004
height_act = 0.0008
height = height_pas + height_act
n_Elem_y_pas = 2
n_Elem_y_act = 4
n_Elem_y = n_Elem_y_pas + n_Elem_y_act
n_Elem_x = 75
meshsize_y = (height_pas + height_act) / n_Elem_y
meshsize_x = length / n_Elem_x
c_dummy = 10000
n_Node_x = n_Elem_x + 1
n_Node_y = n_Elem_y + 1


def node():
    '''
    Define nodes coordinates for the active layer and the pasive layer

    **Parameters**
        global variables

    **Returns**
        node: *dict*
        print in .txt file
    '''
    coor_y = - meshsize_y
    node = dict()
    n_Node = 0
    # Define the coordinate of nodes based on the row and column of the mesh
    for row_Node in range(1, n_Node_y + 1):
        coor_x = - meshsize_x
        coor_y += meshsize_y
        column_Node = 0
        for column_Node in range(1, n_Node_x + 1):
            n_Node += 1
            coor_x += meshsize_x
            node.update({n_Node: (coor_x, coor_x)})
            print(n_Node, ', ', coor_x, ', ', coor_y, sep='', file=f)
    return node


def elem():
    '''
    Define elements for the passive layer and the active layer

    **Parameters**
        global variables

    **Returns**
        elem_pas, elem_act: *dict*
        print in .txt file
    '''

    # Print the heading for element nodes of the passive layer
    print('**\
         \n** The following are elements for the passive layer\
         \n**\
         \n*Element, type=CPE4H', file=f)
    n_Elem = 0

    # Define and print the element nodes of the passive layer based on the row
    # and column of the element
    elem_pas = dict()
    for row_Elem in range(1, n_Elem_y_pas + 1):
        for column_Elem in range(1, n_Elem_x + 1):
            n_Elem += 1
            Node_left_bot = (row_Elem - 1) * n_Node_x + column_Elem
            Node_right_bot = Node_left_bot + 1
            Node_left_top = row_Elem * n_Node_x + column_Elem
            Node_right_top = Node_left_top + 1
            elem_pas.update({n_Elem: (Node_left_bot, Node_right_bot,
                                      Node_right_top, Node_left_top)})
            print(n_Elem, ', ', Node_left_bot, ', ', Node_right_bot,
                  ', ', Node_right_top, ', ', Node_left_top, sep='', file=f)
    print('*Nset, nset=PassiveSet, generate\
         \n%d, %d, 1' % (list(elem_pas.values())[0][0],
                         list(elem_pas.values())[-1][-2]), file=f)
    print('*Elset, elset=PassiveSet, generate\
         \n%d, %d, 1' % (list(elem_pas.keys())[0],
                         list(elem_pas.keys())[-1]), file=f)

    # Print the heading for element nodes of the active layer
    print('**\
         \n** The following aare the UEL elements for the active layer\
         \n**\
         \n*User Element,Nodes=4,Type=U1,Iproperties=2,Properties=14,\
            Coordinates=2,Variables=4,Unsymm\
         \n 1,2,11,12\
         \n*Element, type=U1', file=f)

    # Define and print the nodes of each element of the active layer based
    # on the row and column of the element
    elem_act = dict()
    for row_Elem in range(n_Elem_y_pas + 1, n_Elem_y + 1):
        for column_Elem in range(1, n_Elem_x + 1):
            n_Elem += 1
            Node_left_bot = (row_Elem - 1) * n_Node_x + column_Elem
            Node_right_bot = Node_left_bot + 1
            Node_left_top = row_Elem * n_Node_x + column_Elem
            Node_right_top = Node_left_top + 1
            elem_act.update({n_Elem: (Node_left_bot, Node_right_bot,
                                      Node_right_top, Node_left_top)})
            print(n_Elem, ', ', Node_left_bot, ', ', Node_right_bot,
                  ', ', Node_right_top, ', ', Node_left_top, sep='', file=f)
    print('*Nset, nset=ActiveSet, generate\
         \n%d, %d, 1' % (list(elem_act.values())[0][0],
                         list(elem_act.values())[-1][-2]), file=f)
    print('*Elset, elset=ActiveSet, generate\
         \n%d, %d, 1' % (list(elem_act.keys())[0],
                         list(elem_act.keys())[-1]), file=f)
    return elem_pas, elem_act


def dummmy_elem(elem_act):
    '''
    Define dummy elements

    **Parameters**
        global variables
        elem_act: *dict*

    **Returns**
        print in .txt file
    '''
    # Print the heading for the dummy elements
    print('**\
         \n** Make the dummy mesh used for visualization. These dummy elements\
         \n** use the same nodes as the real mesh, but note the offset in\
         \n** numbering of %d, that shows up again in the UVARM subroutine.\
         \n**\
         \n**Element, type=CPE4' % c_dummy, file=f)

    # Print the dummy elements nodes by an offset of c_dummy
    for key in elem_act:
        print(key + c_dummy, ', ',
              str(elem_act[key])[1:-1], sep='', file=f)
    print('*Elset, elset=elDummy, generate\
         \n%d, %d, 1' % (list(elem_act.keys())[0] + c_dummy,
                         list(elem_act.keys())[-1] + c_dummy), file=f)


def sets_for_BC(node):
    '''
    Define node sets and element sets for boundary conditions

    **Parameters**
        node: *dict*

    **Return**
        print in .txt file
    '''
    print('**\
         \n** Sets used for BCs\
         \n**', file=f)

    print('*Nset, nset=Nall, generate\
         \n %d, %d, 1' % (list(node.keys())[0], list(node.keys())[-1]), file=f)

    print('*Nset, nset=Top, generate\
         \n %d, %d, 1' % (list(node.keys())[-n_Node_x],
                          list(node.keys())[-1]), file=f)
    print('*Elset, elset=Top, generate\
         \n %d, %d, 1' % (list(elem_act.keys())[-n_Elem_x],
                          list(elem_act.keys())[-1]), file=f)

    print('*Nset, nset=Bottom, generate\
         \n %d, %d, 1' % (list(node.keys())[0],
                          list(node.keys())[n_Node_x - 1]), file=f)
    print('*Elset, elset=Bottom, generate\
         \n %d, %d, 1' % (list(elem_pas.keys())[0],
                          list(elem_pas.keys())[n_Elem_x - 1]), file=f)

    print('*Nset, nset=Left, generate\
         \n %d, %d, %d' % (list(node.keys())[0],
                           list(node.keys())[-n_Node_x], n_Node_x), file=f)
    print('*Nset, nset=Left1, generate\
         \n %d, %d, %d' % (list(node.keys())[n_Node_x],
                           list(node.keys())[-2 * n_Node_x],
                           n_Node_x), file=f)
    print('*Elset, elset=Left, generate\
         \n %d, %d, %d' % (list(elem_pas.keys())[0],
                           list(elem_act.keys())[-n_Elem_x],
                           n_Elem_x), file=f)

    print('*Nset, nset=Right, generate\
         \n %d, %d, %d' % (list(node.keys())[n_Node_x - 1],
                           list(node.keys())[-1], n_Node_x), file=f)
    print('*Nset, nset=Right1, generate\
         \n %d, %d, %d' % (list(node.keys())[n_Node_x * 2 - 1],
                           list(node.keys())[-n_Node_x - 1], n_Node_x), file=f)
    print('*Elset, elset=Right, generate\
         \n %d, %d, %d' % (list(elem_pas.keys())[n_Elem_x - 1],
                           list(elem_act.keys())[-1], n_Elem_x), file=f)

    print('*Nset, nset=n1\
        \n %d' % list(node.keys())[n_Node_x - 1], file=f)
    print('*Nset, nset=n2\
        \n %d' % list(node.keys())[-1], file=f)
    print('*Nset, nset=n3\
        \n %d' % list(node.keys())[-n_Node_x], file=f)
    print('*Nset, nset=n4\
        \n %d' % list(node.keys())[0], file=f)


def draw_BC():
    '''
    Output the certain parameters in the parameter file to an excel file

    **Parameters**
        parameter_file: .txt

    **Return**
        excel file: output_parameters.xlsx
    '''
    # Create an image
    im = Image.new('RGB', (2000, 2000), (225, 225, 225))
    draw = ImageDraw.Draw(im)
    scale = 100000
    # Define the corner coordinates of the bilayer part
    rect_x_start = 200
    rect_y_start = 1000
    rect_x_end = rect_x_start + scale * length
    rect_y_end = rect_y_start + scale * height

    # Define the corner coordinates of the active layer
    rect_act = (rect_x_start, rect_y_start, rect_x_start + scale * length,
                rect_y_start + scale * height_act)
    draw.rectangle(rect_act, outline=(0, 0, 0), width=5)

    # Define the corner coordinates of the passive layer
    rect_pas = (rect_x_start, rect_y_start + scale * height_act,
                rect_x_start + scale * length,
                rect_y_start + scale * height_act + scale * height_pas)
    draw.rectangle(rect_pas, outline=(0, 0, 0), width=5)

    # Draw the rigid surface below the bilayer part
    extend = 100
    draw.line((rect_x_start - extend, rect_y_end, rect_x_end + extend,
               rect_y_end), fill=(0, 0, 255), width=5)

    # Draw the boundary condition at both sides of the bilayer part
    tri = 20
    draw.polygon([(rect_x_start, rect_y_end), (rect_x_start - tri,
                                               rect_y_end + tri * 2),
                  (rect_x_start + tri,
                   rect_y_end + tri * 2)],
                 fill=(255, 0, 0))
    draw.polygon([(rect_x_end, rect_y_end), (rect_x_end - tri,
                                             rect_y_end + tri * 2),
                  (rect_x_end + tri,
                   rect_y_end + tri * 2)],
                 fill=(255, 0, 0))

    im.save('output_BC.jpg')


def parameter(parameters_file):
    '''
    Output the certain parameters in the parameter file to an excel file

    **Parameters**
        parameter_file: .txt

    **Return**
        excel file: output_parameters.xlsx
        parameter: *dict*
    '''
    # Define the parameter list that will be output in the excel file
    parameter_list = ['phi0 =', 'theta0 = ', 'thetaHot =', 'initMU =']

    # Create excel file
    f_input = open(parameters_file, 'r')
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'], sheet['B1'] = 'Parameters', 'Values'

    # Define styles of cells
    bold_font = Font(bold=True)
    grey_fill = PatternFill(fgColor='C0C0C0', fill_type='solid')
    for cell in sheet["1:1"]:
        cell.font = bold_font
        cell.fill = grey_fill

    # Save the parameters in a dict
    parameter = dict()
    for line in f_input:
        if any(param in line for param in parameter_list):
            parameter.update(
                {list(line.split('='))[0]: list(line.split('='))[-1]})
    # Input the paramter names and values into the excel
    row_cell = 2
    column_cell = 1
    for key in parameter:
        sheet.cell(row=row_cell, column=column_cell).value = key
        sheet.cell(row=row_cell, column=column_cell + 1).value = parameter[key]
        row_cell += 1

    f_input.close()
    workbook.save(filename="output_parameters.xlsx")

    return parameter


if __name__ == "__main__":
    # Output an txt file for geometry information
    f = open('output_geometry.txt', 'w')
    geometry_file = 'output_geometry.txt'
    node = node()
    elem_pas, elem_act = elem()
    dummmy_elem(elem_act)
    sets_for_BC(node)
    f.close()

    # Output an image file for boundary conditions
    draw_BC()

    # Output an excel file for parameters in the file
    parameter('Parameter.txt')
